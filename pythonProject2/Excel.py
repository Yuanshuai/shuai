
from openpyxl import Workbook
#工具类
from openpyxl.styles import Alignment
import os





class ExcelUtils:
    def __init__(self,filename=None):
        # 直接创建一个新的工作簿
        self.wb = Workbook()
        # 获取活动工作表
        self.ws = self.wb.active
        # 如果提供了文件名，则保存文件名以供后续保存工作簿使用
        if filename:
            self.filename = filename
        else:
            # 如果没有提供文件名，则设置一个默认文件名
            self.filename = 'default.xlsx'
    @staticmethod
    def excel_column_name(index):
        result = ""
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result

    @staticmethod
    def get_cell_address(row_number, column_index):
        column_name = ExcelUtils.excel_column_name(column_index)
        return f"{column_name}{row_number}"

    def hebing(self,detail):
        kaishix=None
        kaishiy=None
        jieshx=None
        jieshuy=None
        newtext=""
        for celly in range(1,len(detail)):
            g0=None
            gl=None
            for cellx in range(1,len(detail[0])):
                shang,xia,zuo,you=detail[celly][cellx]
                if g0==None and not shang:
                        kaishiy=celly
                        kaishix=cellx
                        g0=ExcelUtils.get_cell_address(kaishiy,kaishix)

                elif g0!=None and (cellx==len(detail[0])-1 or shang) :
                        jieshuy=celly
                        jieshux=cellx
                        if jieshux-kaishix<=1:
                            g0 = None
                            gl = None
                        else:
                            gl=ExcelUtils.get_cell_address(jieshuy,jieshux)
                            for g in range(kaishix,jieshux):
                               gg=ExcelUtils.get_cell_address(kaishiy,g)
                               if self.ws[gg].value!=None:
                                    newtext=newtext+self.ws[gg].value

                            self.ws[g0]=newtext
                            self.ws.merge_cells(f"{g0}:{gl}")
                            # 设置合并单元格的居中对齐
                            self.ws[g0].alignment = Alignment(horizontal='center', vertical='center')
                            g0=None
                            gl=None
                            newtext=""
        for cellx in range(1,len(detail[0])):
            g0=None
            gl=None
            for celly in range(1,len(detail)):
                shang,xia,zuo,you=detail[celly][cellx]
                if g0==None and not zuo:
                        kaishiy=celly
                        kaishix=cellx
                        g0=ExcelUtils.get_cell_address(kaishiy,kaishix)

                elif g0!=None and (celly==len(detail)-1 or zuo):
                        jieshuy=celly
                        jieshux=cellx
                        if jieshuy-kaishiy<=1:
                            g0 = None
                            gl = None
                        else:
                            gl=ExcelUtils.get_cell_address(jieshuy,jieshux)
                            for g in range(kaishiy,jieshuy):
                               gg=ExcelUtils.get_cell_address(g,kaishix)
                               if self.ws[gg].value!=None:
                                    newtext=newtext+self.ws[gg].value

                            self.ws[g0]=newtext
                            self.ws.merge_cells(f"{g0}:{gl}")
                            self.ws[g0].alignment = Alignment(horizontal='center', vertical='center')
                            g0=None
                            gl=None
                            newtext=""









    def create_excel_table(self,results, ocrtext, detail,pattern=True):



        # 尝试加载现有的工作簿，如果不存在则创建一个新的工作簿
        try:

            # 清空工作表中的
            for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row):
                for cell in row:
                    cell.value = None
        except FileNotFoundError:
            # 文件不存在，创建一个新的工作簿
            wb = Workbook()
            ws = wb.active

        # 将文本识别和直线检测的结果写入Excel文件

        for cellx in range(len(results[0])):
            kuan=0
            for celly in range(len(results)):
                    (x0, y0), (x1, y1), ct = results[celly][cellx]
                    y_,x_=ct
                    clocation=ExcelUtils.get_cell_address(y_,x_)
                    for lin in ocrtext:
                        tlocation, text = lin
                        xx, yy = tlocation
                        if xx > x0 and xx < x1 and yy > y0 and yy < y1:
                            if self.ws[clocation].value!=None:
                                current_value = self.ws[clocation].value
                                new_content = f"{current_value}{text}"
                            else:
                                new_content=text
                            self.ws[clocation].value = new_content
                            lie = ExcelUtils.excel_column_name(x_)

                            if len(new_content)>kuan:
                                kuan=len(new_content)
                                self.ws.column_dimensions[lie].width = int(kuan*2)+5
        if pattern==True:
            ExcelUtils.hebing(self,detail)
        # 保存工作簿
        self.wb.save(filename=self.filename)
        print(f"文件 '{self.filename}' 已创建并保存。")
        # 获取绝对路径
        absolute_path = os.path.abspath(f"{self.filename}")
        # print(absolute_path)
        os.startfile(absolute_path)