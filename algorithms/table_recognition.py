"""表格结构识别与 Excel 导出模块，连接检测结果与最终表格数据。"""

from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import cv2
import numpy as np
import os
import time

from algorithms.table_detect import image2tables, get_table_structure


@dataclass
class TableCell:
    row: int
    col: int
    bbox: List[int]
    text: str = ""
    rowspan: int = 1
    colspan: int = 1


@dataclass
class TableResult:
    rows: int
    cols: int
    cells: List[TableCell]
    grid: List[List[str]]
    header: List[str]
    first_row: List[str]

    def to_meta(self) -> Dict[str, Any]:
        return {
            "rows": self.rows,
            "cols": self.cols,
            "cells": [{"row": c.row, "col": c.col, "bbox": c.bbox, "text": c.text,
                      "rowspan": c.rowspan, "colspan": c.colspan} for c in self.cells],
            "preview": {"header": self.header, "first_row": self.first_row},
        }


def parse_table_structure(table_bgr: np.ndarray) -> Dict[str, Any]:
    """解析表格结构，返回行列信息和单元格位置"""
    groups = image2tables(table_bgr)
    if not groups:
        return {"rows": 0, "cols": 0, "cells": []}
    shapes = groups[0]
    cells = get_table_structure(shapes)
    out_cells = []
    max_row = -1
    max_col = -1
    for cell in cells:
        row = int(cell.get("row", -1))
        col = int(cell.get("col", -1))
        rowspan = int(cell.get("rowspan", 1))
        colspan = int(cell.get("colspan", 1))
        bbox_obj = cell.get("bbox")
        if bbox_obj is None:
            continue
        out_cells.append(
            {
                "row": row,
                "col": col,
                "rowspan": rowspan,
                "colspan": colspan,
                "bbox": [int(bbox_obj.x1), int(bbox_obj.y1), int(bbox_obj.x2), int(bbox_obj.y2)],
            }
        )
        max_row = max(max_row, row + rowspan - 1)
        max_col = max(max_col, col + colspan - 1)
    return {"rows": max_row + 1, "cols": max_col + 1, "cells": out_cells}


def _is_blank_image(img: np.ndarray) -> bool:
    """
    检查图像是否为100%空白（所有像素颜色完全一致）
    可以是纯黑、纯白、纯绿等任何纯色
    """
    if img is None or img.size == 0:
        return True
    
    # 检查所有像素是否完全相同
    first_pixel = img[0, 0]
    if len(img.shape) == 3:
        # 彩色图像：检查所有像素是否等于第一个像素
        return np.all(img == first_pixel)
    else:
        # 灰度图像
        return np.all(img == first_pixel)


def _compute_y_projection(gray: np.ndarray, threshold: int = None) -> np.ndarray:
    """
    计算Y轴投影（Projection Profile）
    对每一行，计算非空白像素的数量
    使用Otsu自动阈值进行二值化
    """
    # 使用Otsu自动阈值进行二值化
    if threshold is None:
        # Otsu自动计算最优阈值
        _, binary = cv2.threshold(gray, 0, 1, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    else:
        # 使用指定阈值
        _, binary = cv2.threshold(gray, threshold, 1, cv2.THRESH_BINARY_INV)

    # 对每一行求和（即非空白像素的数量）
    projection = np.sum(binary, axis=1)
    return projection


def _find_text_lines_by_projection(projection: np.ndarray, img_height: int,
                                   length_tolerance: float = 0.1) -> List[Tuple[int, int]]:
    """
    根据 Y 轴投影判断文本行数

    算法逻辑：
    1. 找到投影的非零区域（内容区域）
    2. 合并距离小于 5 像素的相邻区域
    3. 如果投影集中于一个区域，且该区域占满图片高度的 50% 以上 → 1 行文本
    4. 如果投影集中于 n 个区域，每个区域长度相近（±10%）且长度 > 图像高度/(2n) → n 行文本
    5. 从投影的中间点断开（确保分割点在空白处，不切割文字像素）
    6. 无效投影区域判断：投影区域总数 > 10 或 投影区域长度 > 原始单元格高度的 1/10

    返回：[(start_row, end_row), ...] 文本行区间列表
    """
    h = len(projection)
    if h == 0:
        return []

    # 使用原始单元格图像的高度（而不是粗剪后的高度）
    original_h = img_height

    # ========== 步骤 1：找到所有投影值大于 0 的连续区域 ==========
    content_regions = []
    in_content = False
    start_idx = 0

    for i in range(h):
        if projection[i] > 0 and not in_content:
            in_content = True
            start_idx = i
        elif projection[i] == 0 and in_content:
            in_content = False
            content_regions.append((start_idx, i))

    # 处理到达底部时仍在内容区域的情况
    if in_content:
        content_regions.append((start_idx, h))

    if not content_regions:
        return []

    # ========== 步骤 2：合并距离小于等于3像素的相邻区域 ==========
    MERGE_DISTANCE = 3  # 合并阈值：相邻区域间距小于等于3像素则合并
    merged_regions = [content_regions[0]]  # 初始化，放入第一个区域

    for i in range(1, len(content_regions)):
        prev_start, prev_end = merged_regions[-1]  # 上一个区域（可能是已合并的）
        curr_start, curr_end = content_regions[i]   # 当前区域

        # 计算相邻区域之间的距离
        distance = curr_start - prev_end

        if distance <= MERGE_DISTANCE:
            # 距离小于等于3像素，合并区域
            merged_regions[-1] = (prev_start, curr_end)
        else:
            # 距离大于3像素，保留为独立区域
            merged_regions.append((curr_start, curr_end))

    content_regions = merged_regions

    # ========== 步骤 3：无效投影区域判断 ==========
    # 条件1：投影区域总数 > 10
    if len(content_regions) > 10:
        # 投影区域太多，可能是噪声，返回空列表
        return []

    # 条件2：投影区域长度 > 原始单元格高度的 1/10
    min_region_length = original_h / 10
    valid_content_regions = []
    for start, end in content_regions:
        region_length = end - start
        if region_length > min_region_length:
            valid_content_regions.append((start, end))

    if not valid_content_regions:
        return []

    content_regions = valid_content_regions

    # ========== 步骤 4：计算每个内容区域的长度 ==========
    region_lengths = [end - start for start, end in content_regions]

    # 情况 1：只有一个内容区域，直接返回作为 1 行文本
    if len(content_regions) == 1:
        return content_regions

    # 情况 2：多个内容区域，判断是否满足 n 行文本的条件（使用原始高度）
    n = len(content_regions)
    min_length = original_h / (10 * n)  # 最小长度要求：原始图像高度/(2n)

    # 检查所有区域长度是否都大于最小长度
    valid_regions = []

    for length in region_lengths:
        # 检查长度是否大于最小长度
        if length < min_length:
            # 太短了，不认为是独立的文本行
            continue
        valid_regions.append(True)

    if len(valid_regions) == n:
        # 满足 n 行文本的条件，从投影的中间点断开
        return _split_at_midpoints(projection, content_regions)
    else:
        # 不满足条件，直接返回原始内容区域
        return content_regions


def _compute_x_projection(binary: np.ndarray) -> np.ndarray:
    """
    计算 X 轴投影（Projection Profile）
    对每一列，计算非空白像素的数量
    
    Returns:
        projection: X 轴投影数组，每个元素表示该列的非空白像素数
    """
    projection = np.sum(binary, axis=0)
    return projection


def _find_content_regions_by_projection(projection: np.ndarray) -> List[Tuple[int, int]]:
    """
    根据投影找到所有内容区域
    
    Args:
        projection: 投影数组
        
    Returns:
        [(start, end), ...] 内容区域区间列表
    """
    content_regions = []
    in_content = False
    start_idx = 0
    
    for i in range(len(projection)):
        if projection[i] > 0 and not in_content:
            in_content = True
            start_idx = i
        elif projection[i] == 0 and in_content:
            in_content = False
            content_regions.append((start_idx, i))
    
    # 处理到达末尾时仍在内容区域的情况
    if in_content:
        content_regions.append((start_idx, len(projection)))
    
    return content_regions


def _crop_to_square_with_margin(img: np.ndarray, binary: np.ndarray, margin: int = 10) -> Tuple[List[np.ndarray], List[np.ndarray]]:
    """
    无有效文本时，进行四边裁剪（保留边距），并尽量做成正方形
    1. 找到内容的上下左右边界
    2. 保留margin像素边距
    3. 尽量扩展成正方形（以较长边为基准）

    返回：(裁剪后的原图列表, 对应的二值图列表)
    """
    h, w = binary.shape

    # 找到上下边界（有内容的行）
    top, bottom = 0, h
    for i in range(h):
        if np.sum(binary[i, :]) > 0:
            top = i
            break
    for i in range(h - 1, -1, -1):
        if np.sum(binary[i, :]) > 0:
            bottom = i + 1
            break

    # 找到左右边界（有内容的列）
    left, right = 0, w
    for i in range(w):
        if np.sum(binary[:, i]) > 0:
            left = i
            break
    for i in range(w - 1, -1, -1):
        if np.sum(binary[:, i]) > 0:
            right = i + 1
            break

    # 如果没有内容，返回原图
    if top >= bottom or left >= right:
        return [img], [binary]

    # 计算内容区域的大小
    content_h = bottom - top
    content_w = right - left

    # 以较长边为基准，尽量做成正方形
    square_size = max(content_h, content_w)

    # 计算正方形的中心（基于内容中心）
    center_y = (top + bottom) // 2
    center_x = (left + right) // 2

    # 计算正方形的边界（加上margin）
    half_size = square_size // 2 + margin

    # 确保不超出原图边界
    new_top = max(0, center_y - half_size)
    new_bottom = min(h, center_y + half_size)
    new_left = max(0, center_x - half_size)
    new_right = min(w, center_x + half_size)

    # 如果边界不够，尝试扩展另一边（保持正方形）
    actual_h = new_bottom - new_top
    actual_w = new_right - new_left

    # 裁剪并返回
    cropped = img[new_top:new_bottom, new_left:new_right]
    cropped_binary = binary[new_top:new_bottom, new_left:new_right]
    return [cropped], [cropped_binary]


def _split_at_midpoints(projection: np.ndarray, content_regions: List[Tuple[int, int]]) -> List[Tuple[int, int]]:
    """
    从内容区域的中间点断开
    在相邻内容区域之间的空白区域的中点处断开
    确保分割点只在空白区域（投影为0的地方），不切割任何文字像素
    """
    if len(content_regions) <= 1:
        return content_regions
    
    result = []
    
    # 第一个区域的起点
    result.append((content_regions[0][0], content_regions[0][1]))
    
    # 处理相邻区域之间的分割
    for i in range(1, len(content_regions)):
        prev_end = content_regions[i-1][1]  # 上一个内容区域结束
        curr_start = content_regions[i][0]   # 当前内容区域开始
        
        # 在prev_end和curr_start之间找空白区域的中点
        # 确保分割点在投影为0的位置
        blank_start = prev_end
        blank_end = curr_start
        
        # 找空白区域的实际边界（投影为0的地方）
        # 向前找，确保在空白区域内
        while blank_start < len(projection) and projection[blank_start] > 0:
            blank_start += 1
        # 向后找
        while blank_end > 0 and projection[blank_end - 1] > 0:
            blank_end -= 1
        
        # 在空白区域的中点处断开
        if blank_start < blank_end:
            split_point = (blank_start + blank_end) // 2
        else:
            # 如果没有空白区域，使用原来的边界
            split_point = prev_end
        
        # 确保分割点处投影为0（空白）
        # 如果不在空白处，向两边找最近的空白点
        if projection[split_point] > 0:
            # 向左找空白
            left_blank = split_point
            while left_blank > prev_end and projection[left_blank] > 0:
                left_blank -= 1
            # 向右找空白
            right_blank = split_point
            while right_blank < curr_start and projection[right_blank] > 0:
                right_blank += 1
            
            # 选择离中点最近的空白点
            if projection[left_blank] == 0 and projection[right_blank] == 0:
                if abs(left_blank - split_point) <= abs(right_blank - split_point):
                    split_point = left_blank
                else:
                    split_point = right_blank
            elif projection[left_blank] == 0:
                split_point = left_blank
            elif projection[right_blank] == 0:
                split_point = right_blank
            else:
                # 找不到空白点，使用原边界（不应该发生）
                split_point = prev_end
        
        # 更新上一个区域的结束点
        result[-1] = (result[-1][0], split_point)
        
        # 添加当前区域（从分割点开始）
        result.append((split_point, content_regions[i][1]))
    
    return result


def _crop_blank_edges_v2(img: np.ndarray, threshold: int = None, margin: int = 5) -> Tuple[List[np.ndarray], List[np.ndarray]]:
    """
    基于 Y 轴投影的裁剪逻辑：
    1. 先二值化
    2. 用"剪刀"剪掉四边明显的空白边缘（粗剪）
    3. 计算 Y 轴投影
    4. 根据投影判断文本行数，从中间点断开（精剪）
    5. 再次粗剪到文字边界
    6. X 轴投影分析：如果只有一个区域（单字符），裁剪成正方形；多个区域（多字符）则保留

    返回：(裁剪后的原图列表, 对应的二值图列表)
    """
    if img is None or img.size == 0:
        return [img], [img]

    # 转换为灰度图
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img.copy()
    h, w = gray.shape

    # ========== 步骤0：二值化（自适应阈值）==========
    if threshold is None:
        # 自适应阈值 - 更适合光照不均的文档
        binary = cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,  # 高斯加权平均
            cv2.THRESH_BINARY_INV,  # 反转：文字/线条为白色(255)，背景为黑色(0)
            blockSize=15,  # 邻域大小（奇数）
            C=10  # 常数，从均值中减去
        )
    else:
        # 指定阈值
        _, binary = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY_INV)

    # ========== 步骤0.5：去除细小噪点（基于连通域面积）==========
    # 使用连通域分析，只保留面积大于阈值的连通域
    # 这样可以去除细小的孤立噪点，同时保留文字笔画
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(binary, connectivity=8)

    # 创建新的二值图
    h, w = binary.shape
    binary_clean = np.zeros((h, w), dtype=np.uint8)

    # 面积阈值：小于 10 像素的连通域认为是噪点
    min_area = 10

    for i in range(1, num_labels):  # 跳过背景（标签0）
        area = stats[i, cv2.CC_STAT_AREA]
        if area >= min_area:
            # 保留这个连通域
            binary_clean[labels == i] = 255

    binary = binary_clean

    # 归一化到 0-1 范围
    binary = binary // 255

    # ========== 步骤1：四边粗剪（基于二值图，剪掉明显的空白边缘）==========

    # 1.1 左右裁剪：在二值图上找到第一个和最后一个有内容的列
    first_content_col = 0
    for i in range(w):
        if np.sum(binary[:, i]) > 0:  # 该列有内容（投影>0）
            first_content_col = i
            break

    last_content_col = w
    for i in range(w - 1, -1, -1):
        if np.sum(binary[:, i]) > 0:
            last_content_col = i + 1
            break

    # 确保找到的内容区域有效（有实际内容）
    if first_content_col >= last_content_col:
        return [img]

    # 剪掉左右空白，保留边距（在原图上裁剪）
    left_with_margin = max(0, first_content_col - margin)
    right_with_margin = min(w, last_content_col + margin)
    img_lr_cropped = img[:, left_with_margin:right_with_margin]
    binary_lr_cropped = binary[:, left_with_margin:right_with_margin]

    # 1.2 上下裁剪：在二值图上找到第一个和最后一个有内容的行
    cropped_h, cropped_w = binary_lr_cropped.shape

    first_content_row = 0
    for i in range(cropped_h):
        if np.sum(binary_lr_cropped[i, :]) > 0:
            first_content_row = i
            break

    last_content_row = cropped_h
    for i in range(cropped_h - 1, -1, -1):
        if np.sum(binary_lr_cropped[i, :]) > 0:
            last_content_row = i + 1
            break

    # 确保找到的内容区域有效（有实际内容）
    if first_content_row >= last_content_row:
        return [img]

    # 剪掉上下空白，保留边距（在原图上裁剪）
    top_with_margin = max(0, first_content_row - margin)
    bottom_with_margin = min(cropped_h, last_content_row + margin)

    # 粗剪后的图像（带边距）
    img_cropped = img_lr_cropped[top_with_margin:bottom_with_margin, :]
    binary_cropped = binary_lr_cropped[top_with_margin:bottom_with_margin, :]
    cropped_h_final, cropped_w_final = binary_cropped.shape

    # 记录粗剪的偏移量（用于后续映射回原图）
    offset_x = left_with_margin
    offset_y = top_with_margin

    # ========== 步骤2：计算Y轴投影（基于二值图）==========
    projection = np.sum(binary_cropped, axis=1)

    # ========== 步骤 3：根据投影判断文本行数 ==========
    # 传入原始单元格图像的高度（而不是粗剪后的高度）
    text_lines = _find_text_lines_by_projection(projection, h)

    if not text_lines:
        # 没有检测到有效文本，进行四边裁剪（保留 10 像素边距），并尽量做成正方形
        result_crops, binary_crops = _crop_to_square_with_margin(img, binary, margin=5)
        return result_crops, binary_crops

    # ========== 步骤 4：基于投影结果精剪，映射回原图坐标（加上边距）==========
    result_crops = []
    binary_crops = []
    for idx, (start_row, end_row) in enumerate(text_lines):
        # text_lines 是基于 binary_cropped 的坐标
        # 映射回原图坐标
        base_top = offset_y + start_row
        base_bottom = offset_y + end_row
        
        # 计算边距，确保不进入其他投影区域
        # 上边距：不能超过上一个投影区域的底部
        if idx > 0:
            prev_bottom = offset_y + text_lines[idx - 1][1]
            max_top_margin = base_top - prev_bottom
        else:
            max_top_margin = base_top  # 第一个区域，最多到图像顶部
        top_margin = min(margin, max_top_margin // 2)  # 留一半空间作为边距
        
        # 下边距：不能超过下一个投影区域的顶部
        if idx < len(text_lines) - 1:
            next_top = offset_y + text_lines[idx + 1][0]
            max_bottom_margin = next_top - base_bottom
        else:
            max_bottom_margin = h - base_bottom  # 最后一个区域，最多到图像底部
        bottom_margin = min(margin, max_bottom_margin // 2)  # 留一半空间作为边距
        
        precise_top = max(0, base_top - top_margin)
        precise_bottom = min(h, base_bottom + bottom_margin)
        precise_left = max(0, offset_x - margin)
        precise_right = min(w, offset_x + cropped_w_final + margin)

        # ========== 步骤 5：再次粗剪 - 在二值图上精确裁剪到文字边界（已包含边距）==========
        # 从精剪后的区域中，在二值图上精确找到文字边界
        precise_region = img[precise_top:precise_bottom, precise_left:precise_right]
        precise_binary = binary[precise_top:precise_bottom, precise_left:precise_right]
        precise_h, precise_w = precise_binary.shape

        # 在二值图上找到文字的实际边界（投影>0 的地方）
        # 上下边界
        text_top = 0
        for i in range(precise_h):
            if np.sum(precise_binary[i, :]) > 0:
                text_top = i
                break

        text_bottom = precise_h
        for i in range(precise_h - 1, -1, -1):
            if np.sum(precise_binary[i, :]) > 0:
                text_bottom = i + 1
                break

        # 左右边界
        text_left = 0
        for i in range(precise_w):
            if np.sum(precise_binary[:, i]) > 0:
                text_left = i
                break

        text_right = precise_w
        for i in range(precise_w - 1, -1, -1):
            if np.sum(precise_binary[:, i]) > 0:
                text_right = i + 1
                break

        # 确保不剪掉任何文字像素（扩展边界到包含所有文字）
        # 再加上边距（双重边距保护）
        final_top = max(0, text_top - margin)
        final_bottom = min(precise_h, text_bottom + margin)
        final_left = max(0, text_left - margin)
        final_right = min(precise_w, text_right + margin)

        # 裁剪最终区域
        final_crop = precise_region[final_top:final_bottom, final_left:final_right]

        # 直接保留裁剪后的区域（去掉第6步X轴投影分析）
        result_crops.append(final_crop)

        # 同时保存对应的二值图
        final_binary = precise_binary[final_top:final_bottom, final_left:final_right]
        binary_crops.append(final_binary)

    return result_crops, binary_crops


def _crop_merged_cell(img: np.ndarray, rowspan: int, colspan: int, margin: int = 5) -> Tuple[List[np.ndarray], List[np.ndarray]]:
    """
    专门处理合并单元格的裁剪（激进分割版本）

    处理流程：
    1. 先进行上下左右分割，排除掉完全空白的子单元格区域
    2. 对裁剪后的有效区域进行二值化
    3. 使用基于裁剪后区域高度的 1/10 作为无效投影标准
    4. 进行投影分割，分离多行文本

    Args:
        img: 合并单元格的图像
        rowspan: 跨行数
        colspan: 跨列数
        margin: 边距

    Returns:
        (裁剪后的原图列表, 对应的二值图列表)
    """
    if img is None or img.size == 0:
        return [img], [img]

    # 转换为灰度图（用于初步检测文字位置）
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img.copy()
    h, w = gray.shape

    # 锐化处理：使用Unsharp Masking增强文字边缘
    # 1. 高斯模糊
    gaussian = cv2.GaussianBlur(gray, (0, 0), 3)
    # 2. Unsharp Masking: 原图 + (原图 - 模糊) = 2*原图 - 模糊
    gray_sharpened = cv2.addWeighted(gray, 1.5, gaussian, -0.5, 0)
    # 裁剪到有效范围
    gray_sharpened = np.clip(gray_sharpened, 0, 255).astype(np.uint8)

    # 初步二值化（用于检测哪些子单元格有文字）- 使用锐化后的图像
    binary_pre = cv2.adaptiveThreshold(
        gray_sharpened, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        blockSize=15,
        C=10
    )

    # 去噪
    kernel_small = np.ones((2, 2), np.uint8)
    binary_pre = cv2.morphologyEx(binary_pre, cv2.MORPH_OPEN, kernel_small)

    # 计算每个"子单元格"的高度和宽度
    sub_h = h // rowspan
    sub_w = w // colspan

    # ========== 步骤1：上下左右分割，排除空白子单元格 ==========
    # 上下分割：找到第一个和最后一个有文字的子单元格
    first_text_row = 0
    last_text_row = h

    for i in range(rowspan):
        row_start = i * sub_h
        row_end = min((i + 1) * sub_h, h)
        if np.sum(binary_pre[row_start:row_end, :]) > 0:
            first_text_row = row_start
            break

    for i in range(rowspan - 1, -1, -1):
        row_start = i * sub_h
        row_end = min((i + 1) * sub_h, h)
        if np.sum(binary_pre[row_start:row_end, :]) > 0:
            last_text_row = row_end
            break

    # 左右分割：找到第一个和最后一个有文字的子单元格
    first_text_col = 0
    last_text_col = w

    for i in range(colspan):
        col_start = i * sub_w
        col_end = min((i + 1) * sub_w, w)
        if np.sum(binary_pre[:, col_start:col_end]) > 0:
            first_text_col = col_start
            break

    for i in range(colspan - 1, -1, -1):
        col_start = i * sub_w
        col_end = min((i + 1) * sub_w, w)
        if np.sum(binary_pre[:, col_start:col_end]) > 0:
            last_text_col = col_end
            break

    # 裁剪到有效区域（排除空白子单元格）
    effective_top = max(0, first_text_row - margin)
    effective_bottom = min(h, last_text_row + margin)
    effective_left = max(0, first_text_col - margin)
    effective_right = min(w, last_text_col + margin)

    # 提取有效区域
    effective_img = img[effective_top:effective_bottom, effective_left:effective_right]
    effective_h = effective_bottom - effective_top

    if effective_h <= 0:
        return [img]

    # ========== 步骤2：对有效区域进行二值化 ==========
    effective_gray = cv2.cvtColor(effective_img, cv2.COLOR_BGR2GRAY) if len(effective_img.shape) == 3 else effective_img.copy()

    # 锐化处理：使用Unsharp Masking增强文字边缘
    gaussian_eff = cv2.GaussianBlur(effective_gray, (0, 0), 3)
    effective_gray_sharpened = cv2.addWeighted(effective_gray, 1.5, gaussian_eff, -0.5, 0)
    effective_gray_sharpened = np.clip(effective_gray_sharpened, 0, 255).astype(np.uint8)

    effective_binary = cv2.adaptiveThreshold(
        effective_gray_sharpened, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV,
        blockSize=15,
        C=10
    )

    # 去噪
    effective_binary = cv2.morphologyEx(effective_binary, cv2.MORPH_OPEN, kernel_small)

    # 归一化到 0-1 范围
    effective_binary_norm = effective_binary // 255

    # ========== 步骤3：基于有效区域高度计算投影分割 ==========
    # 使用基于有效区域高度的 1/10 作为无效投影标准
    min_region_length = effective_h / 10

    # 计算Y轴投影
    projection = np.sum(effective_binary_norm, axis=1)

    # 找到投影的非零区域
    content_regions = []
    in_content = False
    start_idx = 0

    for i in range(effective_h):
        if projection[i] > 0 and not in_content:
            in_content = True
            start_idx = i
        elif projection[i] == 0 and in_content:
            in_content = False
            content_regions.append((start_idx, i))

    if in_content:
        content_regions.append((start_idx, effective_h))

    # 过滤无效区域（基于有效区域高度的 1/10）
    valid_regions = []
    for start, end in content_regions:
        if (end - start) >= min_region_length:
            valid_regions.append((start, end))

    # 如果没有有效区域，返回整个有效区域
    if not valid_regions:
        return [effective_img], [effective_binary]

    # 如果只有一个有效区域，返回整个有效区域
    if len(valid_regions) == 1:
        return [effective_img], [effective_binary]

    # ========== 步骤4：多行文本分割 ==========
    # 在相邻区域之间的空白处分割
    result_crops = []
    binary_crops = []

    for i, (start_row, end_row) in enumerate(valid_regions):
        # 添加上下边距
        if i == 0:
            # 第一个区域，上边可以扩展到顶部
            crop_top = max(0, start_row - margin)
        else:
            # 非第一个区域，上边距限制在上一区域和当前区域中间
            prev_end = valid_regions[i-1][1]
            crop_top = max(prev_end + (start_row - prev_end) // 2, start_row - margin)
            crop_top = max(crop_top, prev_end)  # 确保不重叠

        if i == len(valid_regions) - 1:
            # 最后一个区域，下边可以扩展到底部
            crop_bottom = min(effective_h, end_row + margin)
        else:
            # 非最后一个区域，下边距限制在当前区域和下一区域中间
            next_start = valid_regions[i+1][0]
            crop_bottom = min(end_row + (next_start - end_row) // 2, end_row + margin)
            crop_bottom = min(crop_bottom, next_start)  # 确保不重叠

        # 裁剪（使用整个宽度，因为左右已经裁剪过了）
        crop = effective_img[crop_top:crop_bottom, :]
        crop_binary = effective_binary[crop_top:crop_bottom, :]

        if crop.size > 0:
            result_crops.append(crop)
            binary_crops.append(crop_binary)

    return (result_crops if result_crops else [effective_img]), (binary_crops if binary_crops else [effective_binary])


def _resize_min_height(img: np.ndarray, min_height: int = 100) -> np.ndarray:
    """调整图像大小，确保最小高度（用于OCR识别），使用整数比例"""
    h, w = img.shape[:2]
    if h >= min_height:
        return img

    # 计算缩放比例
    scale_num = min_height
    scale_den = h

    # 简化分数
    import math
    gcd = math.gcd(scale_num, scale_den)
    scale_num //= gcd
    scale_den //= gcd

    new_h = min_height
    new_w = (w * scale_num) // scale_den
    new_w = max(1, new_w)

    # 关键修改：使用 INTER_NEAREST 代替 INTER_LINEAR
    # 最近邻插值不产生灰度过渡，保持边缘锐利
    return cv2.resize(img, (new_w, new_h), interpolation=cv2.INTER_NEAREST)


def build_table_grid(
    table_bgr: np.ndarray,
    structure_meta: Dict[str, Any],
    rec_ocr_fn,
    min_text_height: int = 32,
) -> TableResult:
    """
    构建表格内容网格
    直接对每个单元格使用rec模型识别文字（不使用det模型检测文本位置）
    """
    rows = int(structure_meta.get("rows") or 0)
    cols = int(structure_meta.get("cols") or 0)
    cells_meta = structure_meta.get("cells") if isinstance(structure_meta.get("cells"), list) else []
    if rows <= 0 or cols <= 0 or not cells_meta:
        empty = []
        return TableResult(rows=0, cols=0, cells=[], grid=[], header=empty, first_row=empty)

    h, w = table_bgr.shape[:2]
    grid: List[List[str]] = [["" for _ in range(cols)] for _ in range(rows)]
    cells: List[TableCell] = []

    for meta in cells_meta:
        try:
            r = int(meta.get("row", -1))
            c = int(meta.get("col", -1))
            rowspan = int(meta.get("rowspan", 1))
            colspan = int(meta.get("colspan", 1))
            if r < 0 or c < 0 or r >= rows or c >= cols:
                continue
            bx = meta.get("bbox") or []
            x1, y1, x2, y2 = [int(v) for v in bx[:4]]
            x1 = max(0, min(x1, w - 1))
            y1 = max(0, min(y1, h - 1))
            x2 = max(0, min(x2, w))
            y2 = max(0, min(y2, h))
            if x2 <= x1 or y2 <= y1:
                continue

            # 裁剪单元格图像
            crop = table_bgr[y1:y2, x1:x2]
            original_h, original_w = crop.shape[:2]

            # 检查是否为纯空白单元格（100%纯色）
            if _is_blank_image(crop):
                text = ""
            else:
                # 判断是否为合并单元格（跨越多行或多列）
                is_merged_cell = (rowspan > 1) or (colspan > 1)

                if is_merged_cell:
                    # 合并单元格：使用特殊的上下左右分割
                    crop_list, binary_list = _crop_merged_cell(crop, rowspan, colspan)
                else:
                    # 普通单元格：使用标准的边缘裁剪
                    crop_list, binary_list = _crop_blank_edges_v2(crop)

                # 对每个裁剪后的子图像进行OCR
                texts = []
                for i, sub_crop in enumerate(crop_list):
                    # 转换为灰度图
                    if len(sub_crop.shape) == 3:
                        sub_crop_gray = cv2.cvtColor(sub_crop, cv2.COLOR_BGR2GRAY)
                    else:
                        sub_crop_gray = sub_crop.copy()

                    # 调整大小确保最小高度100（OCR需要一定高度才能识别）
                    sub_crop_resized = _resize_min_height(sub_crop_gray, min_text_height)

                    # 使用rec模型直接识别文字（传入灰度图）
                    sub_text = str(rec_ocr_fn(sub_crop_resized) or "").strip()
                    sub_text = sub_text.replace("\r", " ").replace("\n", " ")
                    texts.append(sub_text)

                # 合并所有部分的识别结果（从上到下）
                text = "".join(texts)

            grid[r][c] = text
            cells.append(TableCell(row=r, col=c, bbox=[x1, y1, x2, y2], text=text,
                                   rowspan=rowspan, colspan=colspan))
        except Exception as e:
            continue

    header = grid[0] if len(grid) >= 1 else []
    first_row = grid[1] if len(grid) >= 2 else []
    return TableResult(rows=rows, cols=cols, cells=cells, grid=grid, header=header, first_row=first_row)


def save_tables_to_single_workbook(sheet_tables: List[Tuple[str, TableResult]], out_path: str) -> None:
    """将多个表格保存到单个Excel工作簿的不同工作表中，支持合并单元格"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # 删除默认的sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for name, table in sheet_tables:
        ws = wb.create_sheet(title=name[:31])  # Excel工作表名称最大31字符
        
        # 创建占用矩阵，记录哪些单元格已被合并单元格占用
        # 矩阵大小根据表格实际行列数
        max_row = table.rows if table.rows else len(table.grid)
        max_col = table.cols if table.cols else (len(table.grid[0]) if table.grid else 0)
        occupied = [[False] * (max_col + 1) for _ in range(max_row + 1)]  # +1 因为Excel是1-based
        
        # 写入数据，处理合并单元格
        for r_idx, row_data in enumerate(table.grid):
            excel_row = r_idx + 1  # Excel行号是1-based
            for c_idx, cell_value in enumerate(row_data):
                excel_col = c_idx + 1  # Excel列号是1-based
                
                # 如果这个单元格已被占用（被之前的合并单元格），跳过
                if occupied[excel_row][excel_col]:
                    continue
                
                # 查找对应的单元格信息（rowspan, colspan）
                rowspan = 1
                colspan = 1
                for cell in table.cells:
                    if cell.row == r_idx and cell.col == c_idx:
                        rowspan = cell.rowspan
                        colspan = cell.colspan
                        break
                
                # 写入单元格值
                cell = ws.cell(row=excel_row, column=excel_col, value=cell_value)
                
                # 设置表头样式
                if r_idx == 0 and table.header:
                    cell.font = Font(bold=True)
                
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 处理合并单元格
                if rowspan > 1 or colspan > 1:
                    # 计算合并范围
                    start_cell = f"{get_column_letter(excel_col)}{excel_row}"
                    end_cell = f"{get_column_letter(excel_col + colspan - 1)}{excel_row + rowspan - 1}"
                    
                    # 执行合并
                    ws.merge_cells(f"{start_cell}:{end_cell}")
                    
                    # 标记占用的单元格
                    for r in range(excel_row, excel_row + rowspan):
                        for c in range(excel_col, excel_col + colspan):
                            if r <= max_row and c <= max_col:
                                occupied[r][c] = True
        
        # 自动调整列宽
        for col_idx in range(1, max_col + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row_idx in range(1, max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # 设置列宽（根据内容长度 + 2个字符的边距）
            adjusted_width = min(max_length * 2 + 2, 50)  # 最大宽度限制为50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(out_path)
